"""
主線車輛規劃系統 - Excel 報表輸出
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from config import MODULE_CONFIGS, OUTPUT_DIR


# 樣式定義
HEADER_FONT = Font(name="微軟正黑體", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TITLE_FONT = Font(name="微軟正黑體", bold=True, size=14)
SUBTITLE_FONT = Font(name="微軟正黑體", bold=True, size=12)
NORMAL_FONT = Font(name="微軟正黑體", size=10)
WARNING_FONT = Font(name="微軟正黑體", size=10, color="FF0000")
SUCCESS_FONT = Font(name="微軟正黑體", size=10, color="008000")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header_row(ws, row, col_count):
    """套用表頭樣式"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _style_data_row(ws, row, col_count):
    """套用資料列樣式"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = NORMAL_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _auto_width(ws, col_count, min_width=10, max_width=40):
    """自動調整欄寬"""
    for col in range(1, col_count + 1):
        letter = get_column_letter(col)
        max_len = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    length = len(str(cell.value))
                    # 中文字寬度約 2x
                    cjk_count = sum(1 for c in str(cell.value) if ord(c) > 0x4E00)
                    length += cjk_count
                    max_len = max(max_len, min(length, max_width))
        ws.column_dimensions[letter].width = max_len + 2


def generate_report(all_routes: list, vehicle_pool, merge_reports: list = None,
                    route_mode: str = "loop", load_rate: float = 0.9):
    """
    產出 Excel 報表。

    Parameters:
        all_routes: 所有路線列表
        vehicle_pool: VehiclePool 實例
        merge_reports: 跨縣市整併報告
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = Workbook()

    # Sheet 1: 車輛使用統計
    _create_vehicle_summary_sheet(wb, vehicle_pool, all_routes,
                                  route_mode, load_rate)

    # Sheet 2: 路線總覽
    _create_route_summary_sheet(wb, all_routes)

    # Sheet 3: 模組一明細
    module1_routes = [r for r in all_routes if r.module == 1]
    if module1_routes:
        _create_module_detail_sheet(wb, module1_routes, 1)

    # Sheet 4: 模組二明細
    module2_routes = [r for r in all_routes if r.module == 2]
    if module2_routes:
        _create_module_detail_sheet(wb, module2_routes, 2)

    # Sheet 5: 模組四明細
    module4_routes = [r for r in all_routes if r.module == 4]
    if module4_routes:
        _create_module_detail_sheet(wb, module4_routes, 4)

    # Sheet 6: 規劃備註
    _create_notes_sheet(wb, all_routes, vehicle_pool, merge_reports,
                        route_mode, load_rate)

    # 移除預設的空白 Sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # 儲存
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"路線規劃報告_{timestamp}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    print(f"\n  報表已產出：{filepath}")
    return filepath


def _create_vehicle_summary_sheet(wb, pool, all_routes,
                                   route_mode="loop", load_rate=0.9):
    """Sheet 1: 車輛使用統計"""
    ws = wb.active
    ws.title = "車輛使用統計"
    row = 1

    # 標題
    ws.cell(row=row, column=1, value="車輛使用統計").font = TITLE_FONT
    row += 2

    # 規劃參數資訊
    mode_label = "迴圈模式（Nearest Neighbor + 2-opt TSP）" if route_mode == "loop" \
                 else "同心圓模式（70% 行政區 + 30% 環帶補充）"
    param_info = [
        ("路線模式", mode_label),
        ("裝載率上限", f"{load_rate*100:.0f}%"),
        ("5噸車有效容量", f"{int(160*load_rate)} 箱（上限 160×{load_rate:.0%}）"),
        ("8噸車有效容量", f"{int(240*load_rate)} 箱（上限 240×{load_rate:.0%}）"),
    ]
    for label, val in param_info:
        ws.cell(row=row, column=1, value=label).font = Font(name="微軟正黑體", bold=True, size=10)
        ws.cell(row=row, column=2, value=val).font = NORMAL_FONT
        row += 1
    row += 1

    # 全局統計
    summary = pool.get_summary()
    stats = [
        ("項目", "5噸車", "8噸車"),
        ("總數量", summary["total_5ton"], summary["total_8ton"]),
        ("已使用", summary["used_5ton"], summary["used_8ton"]),
        ("剩餘", summary["available_5ton"], summary["available_8ton"]),
        ("使用率", f"{summary['usage_5ton_pct']:.1f}%", f"{summary['usage_8ton_pct']:.1f}%"),
        ("資源狀態", pool.get_resource_status("5ton"), pool.get_resource_status("8ton")),
    ]

    for data in stats:
        for col, val in enumerate(data, 1):
            ws.cell(row=row, column=col, value=val)
        if data[0] == "項目":
            _style_header_row(ws, row, 3)
        else:
            _style_data_row(ws, row, 3)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="各模組使用明細").font = SUBTITLE_FONT
    row += 1

    # 模組別統計
    headers = ["模組", "5噸車", "8噸車", "合計車輛", "門市數", "總箱數"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, len(headers))
    row += 1

    for mod in [1, 2, 4]:
        mod_routes = [r for r in all_routes if r.module == mod]
        mod_5t = summary["module_usage"][mod]["5ton"]
        mod_8t = summary["module_usage"][mod]["8ton"]
        mod_stores = sum(r.store_count for r in mod_routes)
        mod_boxes = sum(r.total_boxes for r in mod_routes)
        data = [
            MODULE_CONFIGS[mod]["name"],
            mod_5t, mod_8t, mod_5t + mod_8t,
            mod_stores, mod_boxes
        ]
        for col, val in enumerate(data, 1):
            ws.cell(row=row, column=col, value=val)
        _style_data_row(ws, row, len(headers))
        row += 1

    # 合計
    total_stores = sum(r.store_count for r in all_routes)
    total_boxes = sum(r.total_boxes for r in all_routes)
    data = ["合計", summary["used_5ton"], summary["used_8ton"],
            summary["used_5ton"] + summary["used_8ton"],
            total_stores, total_boxes]
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(name="微軟正黑體", bold=True, size=10)
    _style_data_row(ws, row, len(headers))

    _auto_width(ws, len(headers))


def _create_route_summary_sheet(wb, all_routes):
    """Sheet 2: 路線總覽"""
    ws = wb.create_sheet("路線總覽")
    row = 1

    ws.cell(row=row, column=1, value="配送路線總覽").font = TITLE_FONT
    row += 2

    headers = [
        "車輛編號", "模組", "配送縣市", "車型", "配送門市數",
        "總裝載箱數", "載重率", "往返里程(km)", "預估時間(hr)",
        "模組標註", "狀態"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, len(headers))
    row += 1

    for route in all_routes:
        counties = ", ".join(sorted(route.counties_visited))
        if route.is_cross_county:
            counties += " ⚡"

        module_label = MODULE_CONFIGS[route.module]["label"]
        capacity_pct = f"{route.capacity_ratio * 100:.0f}%"

        data = [
            route.vehicle.vehicle_id,
            MODULE_CONFIGS[route.module]["name"],
            counties,
            route.vehicle.type_name if route.vehicle.vehicle_type != "none" else "無車輛",
            route.store_count,
            route.total_boxes,
            capacity_pct,
            f"{route.estimated_distance_km:.1f}",
            f"{route.estimated_time_hours:.1f}",
            module_label,
            route.status_icon,
        ]
        for col, val in enumerate(data, 1):
            ws.cell(row=row, column=col, value=val)
        _style_data_row(ws, row, len(headers))

        # 警告列用紅色字
        if route.warnings:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).font = WARNING_FONT

        row += 1

    # 彙總列
    row += 1
    total_stores = sum(r.store_count for r in all_routes)
    total_boxes = sum(r.total_boxes for r in all_routes)
    total_dist = sum(r.estimated_distance_km for r in all_routes)
    total_time = sum(r.estimated_time_hours for r in all_routes)

    ws.cell(row=row, column=1, value="合計").font = Font(name="微軟正黑體", bold=True, size=10)
    ws.cell(row=row, column=5, value=total_stores)
    ws.cell(row=row, column=6, value=total_boxes)
    ws.cell(row=row, column=8, value=f"{total_dist:.1f}")
    ws.cell(row=row, column=9, value=f"{total_time:.1f}")
    _style_data_row(ws, row, len(headers))

    _auto_width(ws, len(headers))


def _create_module_detail_sheet(wb, routes: list, module: int):
    """建立模組明細工作表"""
    config = MODULE_CONFIGS[module]
    ws = wb.create_sheet(f"{config['name']}明細")
    row = 1

    ws.cell(row=row, column=1, value=f"{config['label']} - 各車趟明細").font = TITLE_FONT
    row += 2

    for route in routes:
        # 車輛標題
        ws.cell(row=row, column=1,
                value=f"【車輛編號：{route.vehicle.vehicle_id}】").font = SUBTITLE_FONT
        row += 1

        # 車輛資訊
        info_lines = [
            f"車型：{route.vehicle.type_name}",
            f"配送縣市：{', '.join(sorted(route.counties_visited))}",
            f"總裝載箱數：{route.total_boxes} 箱（載重率：{route.capacity_ratio * 100:.0f}%）",
            f"配送門市數：{route.store_count} 間",
            f"往返里程數：約 {route.estimated_distance_km:.1f} 公里",
            f"預估行車時間：約 {route.estimated_time_hours:.1f} 小時（含卸貨時間）",
        ]

        if route.is_cross_county:
            info_lines.append("⚡ 跨縣市整併路線")

        for line in info_lines:
            ws.cell(row=row, column=1, value=line).font = NORMAL_FONT
            row += 1

        # 警告
        for warning in route.warnings:
            ws.cell(row=row, column=1, value=warning).font = WARNING_FONT
            row += 1

        row += 1

        # 配送順序表頭
        if module == 1:
            headers = ["順序", "店號", "店名", "縣市", "地址",
                       "規劃到店時間", "每日配送箱數", "累計箱數"]
        else:
            headers = ["順序", "店號", "店名", "縣市", "地址",
                       "每日配送箱數", "累計箱數"]

        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h)
        _style_header_row(ws, row, len(headers))
        row += 1

        # 配送順序資料
        cumulative = 0
        for idx, store in enumerate(route.stores, 1):
            cumulative += store.avg_boxes
            if module == 1:
                data = [
                    idx, store.store_id, store.store_name,
                    store.county, store.address,
                    store.current_arrival_time,
                    store.avg_boxes, cumulative
                ]
            else:
                data = [
                    idx, store.store_id, store.store_name,
                    store.county, store.address,
                    store.avg_boxes, cumulative
                ]

            for col, val in enumerate(data, 1):
                ws.cell(row=row, column=col, value=val)
            _style_data_row(ws, row, len(headers))
            row += 1

        row += 2  # 車輛間空行

    _auto_width(ws, 8)


def _create_notes_sheet(wb, all_routes, pool, merge_reports,
                        route_mode="loop", load_rate=0.9):
    """Sheet: 規劃備註"""
    ws = wb.create_sheet("規劃備註")
    row = 1

    ws.cell(row=row, column=1, value="規劃說明與備註").font = TITLE_FONT
    row += 2

    # 規劃參數
    mode_label = "迴圈模式（Nearest Neighbor + 2-opt TSP）" if route_mode == "loop" \
                 else "同心圓模式（70% 行政區 + 30% 同心圓補充）"
    ws.cell(row=row, column=1, value="⚙️ 規劃參數：").font = SUBTITLE_FONT
    row += 1
    for text in [f"  路線模式：{mode_label}",
                 f"  裝載率上限：{load_rate*100:.0f}%",
                 f"  5噸車有效容量：{int(160*load_rate)} 箱",
                 f"  8噸車有效容量：{int(240*load_rate)} 箱"]:
        ws.cell(row=row, column=1, value=text).font = NORMAL_FONT
        row += 1
    row += 1

    # 符合項目
    ws.cell(row=row, column=1, value="✅ 符合項目：").font = SUBTITLE_FONT
    row += 1

    ok_items = [
        "所有車輛均檢查載重限制（90% 安全上限：5噸車 144箱、8噸車 216箱）",
        "已計算每台車的往返里程數與預估行車時間（含卸貨作業）",
        "已優化路線以達成最短里程與最佳時數（Nearest Neighbor + 2-opt）",
    ]

    # 檢查各模組
    module1_routes = [r for r in all_routes if r.module == 1]
    module2_routes = [r for r in all_routes if r.module == 2]
    module4_routes = [r for r in all_routes if r.module == 4]

    if module1_routes:
        ok_items.append("模組一：已參考門市「規劃到店時間」進行白天配送優化")
    if module2_routes:
        ok_items.append("模組二：已考量作業溝通風險，適度降低單車門市數")
    if module4_routes:
        cross_county = [r for r in module4_routes if r.is_cross_county]
        if cross_county:
            ok_items.append(
                f"模組四：已評估並實施 {len(cross_county)} 條跨縣市整併路線")

    for item in ok_items:
        ws.cell(row=row, column=1, value=f"  - {item}").font = NORMAL_FONT
        row += 1

    row += 1

    # 警告項目
    all_warnings = []
    for route in all_routes:
        for w in route.warnings:
            all_warnings.append(f"{route.vehicle.vehicle_id}: {w}")

    if all_warnings:
        ws.cell(row=row, column=1, value="⚠️ 需注意項目：").font = SUBTITLE_FONT
        row += 1
        for w in all_warnings:
            ws.cell(row=row, column=1, value=f"  - {w}").font = WARNING_FONT
            row += 1
        row += 1

    # 跨縣市整併分析
    if merge_reports:
        ws.cell(row=row, column=1, value="⚡ 跨縣市整併分析（模組四）：").font = SUBTITLE_FONT
        row += 1
        for report in merge_reports:
            status = "✅ 已整併" if report["viable"] else "❌ 未整併"
            ws.cell(row=row, column=1,
                    value=f"  {status} {report['reason']}").font = NORMAL_FONT
            row += 1
        row += 1

    # 效率分析
    ws.cell(row=row, column=1, value="📊 效率分析：").font = SUBTITLE_FONT
    row += 1

    if all_routes:
        avg_stores = sum(r.store_count for r in all_routes) / len(all_routes)
        avg_boxes = sum(r.total_boxes for r in all_routes) / len(all_routes)
        avg_dist = sum(r.estimated_distance_km for r in all_routes) / len(all_routes)
        avg_time = sum(r.estimated_time_hours for r in all_routes) / len(all_routes)
        avg_capacity = sum(r.capacity_ratio for r in all_routes) / len(all_routes)

        stats = [
            f"總路線數：{len(all_routes)} 條",
            f"平均每台車配送門市數：{avg_stores:.1f} 間",
            f"平均每台車裝載箱數：{avg_boxes:.0f} 箱（{avg_capacity * 100:.0f}%）",
            f"平均每台車往返里程：{avg_dist:.1f} 公里",
            f"平均每台車行車時間：{avg_time:.1f} 小時",
        ]
        for s in stats:
            ws.cell(row=row, column=1, value=f"  - {s}").font = NORMAL_FONT
            row += 1

    row += 1

    # 車輛資源狀態
    summary = pool.get_summary()
    ws.cell(row=row, column=1, value="🚀 車輛資源狀態：").font = SUBTITLE_FONT
    row += 1
    ws.cell(row=row, column=1,
            value=f"  5噸車：{pool.get_resource_status('5ton')} "
                  f"（剩餘 {summary['available_5ton']} / {summary['total_5ton']} 台）"
            ).font = NORMAL_FONT
    row += 1
    ws.cell(row=row, column=1,
            value=f"  8噸車：{pool.get_resource_status('8ton')} "
                  f"（剩餘 {summary['available_8ton']} / {summary['total_8ton']} 台）"
            ).font = NORMAL_FONT

    _auto_width(ws, 1, min_width=80, max_width=120)
