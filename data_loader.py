"""
主線車輛規劃系統 - Excel 資料讀取
"""
import os
import re
from openpyxl import load_workbook
from models import Store
from box_calculator import calculate_avg_boxes
from config import (
    BASE_DIR, DATA_DIR, INPUT_FILES,
    COL_CATEGORY, COL_ROUTE, COL_ARRIVAL_TIME, COL_STORE_ID,
    COL_STORE_NAME, COL_COUNTY, COL_WAREHOUSE, COL_ADDRESS,
    COL_BOX_START, COL_BOX_END, COUNTY_NORMALIZE,
)


def extract_district(address: str) -> str:
    """從地址中提取行政區名稱"""
    if not address:
        return ""
    # 匹配: 縣市名 + 區/鄉/鎮/市
    match = re.search(r'(.{2,3}[市縣])(.{1,4}[區鄉鎮市])', address)
    if match:
        return match.group(2)
    return ""


def normalize_county(county: str) -> str:
    """標準化縣市名稱"""
    if not county:
        return ""
    county = county.strip()
    return COUNTY_NORMALIZE.get(county, county)


def load_excel_file(filepath: str, module: int) -> list:
    """
    讀取單一 Excel 檔案，回傳 Store 列表。

    Parameters:
        filepath: Excel 檔案完整路徑
        module: 模組編號 (1, 2, 4)
    """
    stores = []
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        # 跳過表頭
        if row_idx == 1:
            continue

        # 確保有足夠欄位
        if not row or len(row) < COL_ADDRESS + 1:
            continue

        # 跳過空白列和樞紐分析表資料
        store_id = row[COL_STORE_ID]
        if store_id is None or str(store_id).strip() == "":
            continue

        # 跳過列標籤/計數列（樞紐分析表尾部）
        category = row[COL_CATEGORY]
        if category is None or str(category).strip() == "":
            # 檢查是否為樞紐分析表資料
            if row[COL_STORE_NAME] is None:
                continue

        store_id = str(store_id).strip()
        store_name = str(row[COL_STORE_NAME] or "").strip()
        county = normalize_county(str(row[COL_COUNTY] or "").strip())
        address = str(row[COL_ADDRESS] or "").strip()
        district = extract_district(address)
        current_route = str(row[COL_ROUTE] or "").strip()
        arrival_time = str(row[COL_ARRIVAL_TIME] or "").strip()

        # 讀取7天出貨箱數
        daily_boxes = []
        max_col = min(COL_BOX_END + 1, len(row))
        for i in range(COL_BOX_START, max_col):
            daily_boxes.append(row[i])
        # 若欄位不足7天，補None
        while len(daily_boxes) < 7:
            daily_boxes.append(None)

        avg_boxes = calculate_avg_boxes(daily_boxes)

        store = Store(
            store_id=store_id,
            store_name=store_name,
            county=county,
            district=district,
            address=address,
            current_route=current_route,
            current_arrival_time=arrival_time,
            daily_boxes=daily_boxes,
            avg_boxes=avg_boxes,
            module=module,
        )
        stores.append(store)

    wb.close()
    return stores


def _data_dir() -> str:
    """回傳資料來源資料夾路徑（優先使用 DATA_DIR，否則用程式所在資料夾）"""
    if DATA_DIR and os.path.isdir(DATA_DIR):
        return DATA_DIR
    return BASE_DIR


def load_all_data() -> dict:
    """
    讀取所有 Excel 檔案，回傳按模組分組的門市資料。

    Returns:
        dict[int, list[Store]]: {1: [...], 2: [...], 4: [...]}
    """
    src_dir = _data_dir()
    stores_by_module = {1: [], 2: [], 4: []}

    # 模組一
    filepath = os.path.join(src_dir, INPUT_FILES["module1"])
    if os.path.exists(filepath):
        stores = load_excel_file(filepath, module=1)
        stores_by_module[1].extend(stores)
        print(f"  模組一：讀取 {len(stores)} 間門市")

    # 模組二
    filepath = os.path.join(src_dir, INPUT_FILES["module2"])
    if os.path.exists(filepath):
        stores = load_excel_file(filepath, module=2)
        stores_by_module[2].extend(stores)
        print(f"  模組二：讀取 {len(stores)} 間門市")

    # 模組四（多個縣市檔案）
    module4_keys = [k for k in INPUT_FILES if k.startswith("module4_")]
    for key in module4_keys:
        filepath = os.path.join(src_dir, INPUT_FILES[key])
        if os.path.exists(filepath):
            stores = load_excel_file(filepath, module=4)
            stores_by_module[4].extend(stores)
            county_name = key.replace("module4_", "")
            print(f"  模組四 ({county_name})：讀取 {len(stores)} 間門市")

    # 彙總
    total = sum(len(v) for v in stores_by_module.values())
    print(f"\n  總計讀取：{total} 間門市")
    print(f"    模組一：{len(stores_by_module[1])} 間")
    print(f"    模組二：{len(stores_by_module[2])} 間")
    print(f"    模組四：{len(stores_by_module[4])} 間")

    return stores_by_module
